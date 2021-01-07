from openpyxl import load_workbook
import getpass

# 불러올 엑셀 파일 설정
# 현재 컴퓨터의 유저명 들고오기
userName = getpass.getuser()

# oneDrive 주소 가져오기
path1 = "C:/Users/"
path2 = "/OneDrive/AI_List.xlsx"

path = path1 + str(userName) + path2

# data_only=True : 수식없이 값만 가져오게 설정
# 내가 저장한 AI_List 들고오기
load_wb = load_workbook(path, data_only=True)

# 불러올 시트 설정
load_sheet = load_wb['Potential']

stock_list = []  # 가능성 종목 리스트
sector_list = []  # 가능성 종목 섹터

# 모든 기록한 종목 가져오기
def call_excell_stock():
    for i in range(5, 100):
        stock_name = load_sheet.cell(i, 3).value
        if stock_name == None:
            break
        stock_list.append(stock_name)

    return stock_list

# 섹터 분류 가져오기
def call_excell_sector():
    for i in range(5, 100):
        sector = load_sheet.cell(i, 4).value
        if sector == None:
            break
        sector_list.append(sector)

    return sector_list