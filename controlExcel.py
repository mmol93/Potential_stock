import datetime

from openpyxl import load_workbook
import getpass

## 불러올 엑셀 파일 설정

# 현재 컴퓨터의 유저명 들고오기
userName = getpass.getuser()

# oneDrive 주소 가져오기
path1 = "C:/Users/"
path2 = "/OneDrive/AI_List.xlsx"

path = path1 + str(userName) + path2

# data_only=True : 수식없이 값만 가져오게 설정
# 내가 저장한 AI_List 들고오기
load_wb = load_workbook(path, data_only=True)

# 불러올 시트 설정
load_sheet = load_wb['매수추천']

StockList = []  # 인공지능 종목 리스트
webPageList = []  # 인공지능 추천가

## *** 엑셀 데이터 저장
def saveExcell(load_wb):
    # 엑셀이 열려있을 때 저장하려고 할 경우 에러 메시지 출력
    try:
        excell_file = load_wb
        excell_file.save(path)
    except PermissionError:
        print("열려있는 AI_List.xlsx 엑셀을 닫으세요")

# 셀에 종목 기록하기
def add_stock(stock, url):
    for i in range(2, 200):
        # 'B2'부터 'B100'까지 읽기
        stock_name = load_sheet.cell(i, 2).value
        # 셀 비었으면 종목명, 해당 종목의 URL 기록하고 함수 종료
        if stock_name == None:
            load_sheet.cell(i, 2).value = stock
            load_sheet.cell(i, 3).value = url
            break
        # 셀에 뭔가 적혀있으면 다음 셀로 이동
        else:
            continue
# 엑셀 저장
    saveExcell(load_wb)

# 기록하기 전에 기존 데이터 삭제하고 시작 날짜 입력하기
def delete_data():
    today = datetime.datetime.now()
    today = today.strftime("%Y-%m-%d")

    # 1, 2, 3번 열 삭제(삭제하면 하나씩 왼쪽으로 오기 때문에 이렇게 3번 삭제한다)
    load_sheet.delete_cols(1)
    load_sheet.delete_cols(1)
    load_sheet.delete_cols(1)

    # 날짜 입력
    load_sheet.cell(1, 2).value = str(today)

    # 엑셀 저장
    saveExcell(load_wb)